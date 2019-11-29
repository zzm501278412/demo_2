from faker import Faker
import xlrd
import os
from openpyxl import Workbook,load_workbook

class Do_Excel:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def write(self, i, j, value):
        if not os.path.exists(self.filename):
            wb = Workbook()
            sh = wb.create_sheet(self.sheetname)
        else:
            wb = load_workbook(self.filename)
            sh = wb[self.sheetname]
        sh.cell(i, j).value = value
        wb.save(self.filename)

if __name__ == '__main__':
    Do_Excel('测试数据1.xlsx', "Sheet1").write(1, 1, 'dep_id')
    Do_Excel('测试数据1.xlsx', "Sheet1").write(1, 2, 'dep_name')
    Do_Excel('测试数据1.xlsx', "Sheet1").write(1, 3, 'master_name')
    Do_Excel('测试数据1.xlsx', "Sheet1").write(1, 4, 'slog')
    # Do_Excel('测试数据1.xlsx', "Sheet1").write(1, 5, 'tel')


    for i in range(2000):
        f = Faker(locale='zh_CN')
        # username = f.user_name()  # 账户名
        email = f.ascii_email()  # 邮箱
        tel = str(f.phone_number())  # 手机号
        a = f.random_int(10000, 9999999)
        b = f.random_letter()
        passsword = '%s%s' % (a, b)
        # passsword1 ='%s%s' % (a, b)
        sousuo = os.path.abspath('测试数据1.xlsx')
        data = xlrd.open_workbook(sousuo)
        sheet1 = data.sheet_by_index(1)
        f = sheet1.nrows + 1
        print(f)
        Do_Excel('测试数据1.xlsx', "Sheet1").write(f, 1, i)
        Do_Excel('测试数据1.xlsx', "Sheet1").write(f, 2, email)
        Do_Excel('测试数据1.xlsx', "Sheet1").write(f, 3,passsword)
        Do_Excel('测试数据1.xlsx', "Sheet1").write(f, 4, tel)
        # Do_Excel('测试数据1.xlsx', "Sheet1").write(f, 5, tel)
        f += 1
